import os
from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
import logging
import uuid
from typing import Optional
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Config
from dotenv import load_dotenv
load_dotenv()

# Services
from services.nlp_service import NLPService
from services.search_service import SearchService

# Schemas
from schemas import SearchRequest, SearchStatus

# Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="ThesisNow Backend", version="0.1.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000", "http://localhost:8000", "https://thesis-now-ai.vercel.app"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Services
try:
    nlp_service = NLPService()
    search_service = SearchService()
except Exception as e:
    logger.error(f"Failed to initialize services: {str(e)}")
    logger.warning("Services will be unavailable until credentials are configured")
    nlp_service = None
    search_service = None

# In-memory job store (TODO: move to Supabase)
jobs = {}

@app.get("/health")
async def health():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}

@app.get("/test/list-jobs")
async def list_jobs():
    """Debug endpoint to list all jobs in memory."""
    job_list = []
    for job_id, job in jobs.items():
        job_list.append({
            "id": job_id,
            "status": job.get("status"),
            "title": job.get("title"),
            "results_count": len(job.get("results", []))
        })
    return {"total": len(jobs), "jobs": job_list}

@app.get("/test/create-sample-job")
async def create_sample_job():
    """
    Create a sample job with test data for Excel export testing.
    """
    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "id": job_id,
        "title": "Impact of Mobile Learning on Academic Performance",
        "status": "completed",
        "boolean_query": '(mobile learning OR m-learning) AND (academic performance OR student achievement)',
        "explanation": "Test query for Excel export",
        "results": [
            {
                "source": "semantic_scholar",
                "title": "Mobile Learning and Student Engagement: A Meta-Analysis",
                "authors": ["Smith, John", "Johnson, Mary"],
                "year": 2023,
                "doi": "10.1234/example.2023",
                "url": "https://example.com/paper1",
                "relevance_score": 0.92,
                "similarity_score": 0.88,
                "citation_count": 15,
                "doc_type": "Journal Article",
                "abstract": "This study examines the impact of mobile learning technologies on student engagement and academic outcomes in higher education."
            },
            {
                "source": "pubmed",
                "title": "Effects of Digital Tools on Learning Outcomes",
                "authors": ["Chen, Li", "Wang, Wei", "Garcia, Carlos"],
                "year": 2022,
                "pmid": "12345678",
                "url": "https://example.com/paper2",
                "relevance_score": 0.85,
                "similarity_score": 0.80,
                "citation_count": 42,
                "doc_type": "Research Paper",
                "abstract": "Digital learning tools have shown significant improvements in student performance across multiple studies."
            },
            {
                "source": "arxiv",
                "title": "Machine Learning for Personalized Learning Paths",
                "authors": ["Kumar, Raj"],
                "year": 2024,
                "doi": "10.5678/arxiv.2024",
                "url": "https://example.com/paper3",
                "relevance_score": 0.72,
                "similarity_score": 0.68,
                "citation_count": 3,
                "doc_type": "Preprint",
                "abstract": "Adaptive learning systems using machine learning can optimize student learning experiences."
            }
        ],
        "created_at": datetime.utcnow().isoformat(),
        "completed_at": datetime.utcnow().isoformat(),
    }
    return {"job_id": job_id, "status": "ready"}

@app.post("/nlp/generate-query")
async def generate_query_only(request: SearchRequest):
    """
    Generate a boolean query without executing the search.
    Used for preview/validation before full search.
    """
    try:
        result = await nlp_service.generate_query_full(request.title, request.lang or "es")
        return {"query": result.get("boolean_query", ""), "explanation": result.get("explanation", "")}
    except Exception as e:
        logger.error(f"Error generating query: {str(e)}")
        return {"query": f'"{request.title}"', "error": str(e)}

@app.post("/search")
async def start_search(request: SearchRequest, background_tasks: BackgroundTasks):
    """
    Start a new bibliographic search.
    Returns immediately with a job_id for polling.
    """
    job_id = str(uuid.uuid4())
    logger.info(f"\n{'='*80}")
    logger.info(f"[{job_id}] NEW SEARCH REQUEST")
    logger.info(f"Title: {request.title}")
    logger.info(f"Databases: {request.databases}")
    logger.info(f"{'='*80}\n")

    jobs[job_id] = {
        "id": job_id,
        "title": request.title,
        "status": "pending",
        "boolean_query": None,
        "explanation": None,
        "results": None,
        "error": None,
        "created_at": datetime.utcnow().isoformat(),
    }

    # Queue the search in background
    background_tasks.add_task(execute_search, job_id, request)

    return {"job_id": job_id, "status": "pending"}

@app.get("/search/{job_id}")
async def get_search_status(job_id: str):
    """
    Poll for search status and results.
    """
    if job_id not in jobs:
        logger.warning(f"[{job_id}] Search not found")
        raise HTTPException(status_code=404, detail="Search not found")

    job = jobs[job_id]
    logger.debug(f"[{job_id}] Status check: {job['status']}")
    return job

@app.get("/report/{job_id}/excel")
async def export_excel(job_id: str):
    """
    Export search results as Excel file.
    """
    if job_id not in jobs:
        logger.warning(f"[{job_id}] Search not found")
        raise HTTPException(status_code=404, detail="Search not found")

    job = jobs[job_id]
    results = job.get("results", [])

    if not results:
        raise HTTPException(status_code=400, detail="No results to export")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    # Headers with styling
    headers = ["Título", "Autor/es", "Año", "Base de datos", "Tipo de estudio"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row_idx, result in enumerate(results, start=2):
        title = result.get("title") if isinstance(result, dict) else result.title or ""
        authors = result.get("authors", []) if isinstance(result, dict) else result.authors or []
        if not authors:
            authors_str = ""
        elif len(authors) == 1:
            authors_str = authors[0]
        else:
            authors_str = f"{authors[0]} et al."
        year = result.get("year") if isinstance(result, dict) else result.year or ""
        source = result.get("source") if isinstance(result, dict) else result.source or ""
        doc_type = result.get("doc_type") if isinstance(result, dict) else result.doc_type or ""

        ws.cell(row=row_idx, column=1, value=title)
        ws.cell(row=row_idx, column=2, value=authors_str)
        ws.cell(row=row_idx, column=3, value=year)
        ws.cell(row=row_idx, column=4, value=source)
        ws.cell(row=row_idx, column=5, value=doc_type)

    # Column widths
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_bytes = excel_file.getvalue()

    filename = f"{job.get('title', 'resultados').replace(' ', '_')}.xlsx"

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

async def execute_search(job_id: str, request: SearchRequest):
    """
    Execute the search in background:
    1. Generate boolean query from title
    2. Search in multiple databases in parallel
    3. Deduplicate and score results
    """
    try:
        # Step 1: NLP - Generate boolean query
        jobs[job_id]["status"] = "generating_query"
        logger.info(f"[{job_id}] ⏳ STEP 1: Generating boolean query...")
        logger.info(f"[{job_id}] Input: {request.title}")

        nlp_result = await nlp_service.generate_query_full(request.title, request.lang or "es")
        jobs[job_id]["boolean_query"] = nlp_result.get("boolean_query", "")
        jobs[job_id]["explanation"] = nlp_result.get("explanation", "")
        logger.info(f"[{job_id}] ✅ Query generated: {nlp_result.get('boolean_query', '')}\n")

        # Step 2: Search
        jobs[job_id]["status"] = "searching"
        databases = request.databases or ["pubmed", "semantic_scholar"]
        logger.info(f"[{job_id}] ⏳ STEP 2: Searching in databases...")
        logger.info(f"[{job_id}] Target databases: {', '.join(databases)}")
        logger.info(f"[{job_id}] Query: {nlp_result.get('boolean_query', '')}\n")

        filters = {
            "year_from": request.year_from,
            "year_to": request.year_to,
            "doc_types": request.doc_types,
            "lang_filter": request.lang_filter,
            "open_access_only": request.open_access_only,
            "peer_reviewed_only": request.peer_reviewed_only,
        }

        results = await search_service.search_all_databases(
            nlp_result.get("boolean_query", ""),
            databases,
            filters
        )

        jobs[job_id]["results"] = results
        jobs[job_id]["status"] = "completed"
        jobs[job_id]["completed_at"] = datetime.utcnow().isoformat()
        logger.info(f"\n[{job_id}] ✅ SEARCH COMPLETED!")
        logger.info(f"[{job_id}] Total results: {len(results)}")
        logger.info(f"[{job_id}] {'='*80}\n")

    except ValueError as e:
        logger.error(f"\n[{job_id}] ❌ ERROR: {str(e)}")
        logger.error(f"[{job_id}] {'='*80}\n")
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"] = str(e)
    except Exception as e:
        logger.error(f"\n[{job_id}] ❌ ERROR: {str(e)}")
        logger.error(f"[{job_id}] {'='*80}\n")
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"] = str(e)

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
